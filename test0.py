import xlrd
import pandas as pd
import openpyxl
path=r'C:\Users\jianya_liao\Desktop\test_package\outdata.xlsx'
file =xlrd.open_workbook(path)
table=file.sheets()[0]
names =file.sheet_names()
print(table)
print('name:',names)


## 行列操作
nrows =table.nrows  # ncols
table_row=table.row(5)  #table.col()
print(nrows)#43
print(table_row)#[text:'F01590', number:5.0, text:'18CSV21B190227C12Q', number:3.057, number:0.004, number:105.098, number:448.5, number:54.0, number:60.0, number:3.18, number:0.0, number:98.41]
print(table.row_types(5))  #array('B', [1, 2, 1, 2, 2, 2, 2, 2, 2, 2, 2, 2])
print(table.row_len(4))  #12
print(table.row_values(3))

#单元格操作
print(table.cell(1,2))  #(行，列)  #t  ext:'18CSV21B190227C12Q'
print(table.cell_type(1,2)) #返回1代表数字，2代表字符
print(table.cell_value(1,2))

wb=openpyxl.load_workbook(r'C:\Users\jianya_liao\Desktop\test_package\outdata.xlsx')
ws=wb.create_sheet('Mysheet1')


row=[1,2,3,4]
ws.append(row)
wb.save('new.xlsx')